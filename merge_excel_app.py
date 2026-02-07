import os
import re
import json
import datetime
import stat
from pathlib import Path
from difflib import SequenceMatcher
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def set_file_readonly(filepath: Path, readonly: bool = True):
    """Set or remove read-only flag on a file."""
    try:
        if readonly:
            # Remove write permissions
            os.chmod(filepath, stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
        else:
            # Add write permissions
            os.chmod(filepath, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
    except Exception:
        pass


def is_file_readonly(filepath: Path) -> bool:
    """Check if a file is read-only (no write permission)."""
    try:
        if not filepath.exists():
            return False
        mode = os.stat(filepath).st_mode
        return not (mode & stat.S_IWUSR)
    except Exception:
        return False


# -------------------------
# Persistent storage paths
# -------------------------
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
ORDERS_DIR = None  # Не се използва - потребителят избира къде да записва
PROTOCOLS_DIR = None  # Ще се зададе от потребителя
SETTINGS_FILE = DATA_DIR / "settings.json"


# -------------------------
# Helpers: header detection
# -------------------------
def normalize(s: str) -> str:
    """Normalize a header or key: lowercase, remove punctuation and collapse whitespace.
    This makes matching more robust for variants like 'шир./вис.' vs 'шир/вис'.
    """
    s = str(s).strip().lower()
    # replace any non-word/digit characters with a single space (keeps Cyrillic/latin letters and digits)
    s = re.sub(r"[^\w\d]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s
def canonical_code(x) -> str:
    """
    Канонизира код/артикул за сравнение между таблици.
    - маха .0 (ако Excel го е прочел като float)
    - маха интервали / NBSP
    - прави upper()
    """
    if pd.isna(x):
        return ""
    # ако е float и е цяло число -> int
    if isinstance(x, float) and x.is_integer():
        x = int(x)
    s = str(x).strip()
    # махни .0 ако идва от float в текст
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    # махни интервали и NBSP
    s = re.sub(r"[\s\u00A0\u202F]+", "", s)
    return s.upper()


def fuzzy_match_best(item: str, candidates: list, threshold: float = 0.6):
    """
    Find the best fuzzy match for `item` among `candidates`.
    Returns (best_candidate, score) or (None, 0) if no match above threshold.
    """
    item_lower = item.lower()
    best = None
    best_score = 0
    for c in candidates:
        if not isinstance(c, str):
            continue
        score = SequenceMatcher(None, item_lower, c.lower()).ratio()
        if score > best_score:
            best_score = score
            best = c
    if best_score >= threshold:
        return best, best_score
    return None, 0


def to_float(x):
    """
    Подобрен float parser:
    - приема '0,85', '0.85', '0,85 лв', '€0.85', '1 234,56'
    - вади първото число от текста
    """
    if pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip()
    if not s:
        return None

    # махни интервали/хилядарни разделители
    s = re.sub(r"[\s\u00A0\u202F]+", "", s)

    # вземи първото число (с , или .)
    m = re.search(r"[-+]?\d+(?:[.,]\d+)?", s)
    if not m:
        return None

    num = m.group(0).replace(",", ".")
    try:
        return float(num)
    except Exception:
        return None


def parse_qty_range_from_header(header: str):
    """
    Извлича (min,max) от заглавия като:
      '1 000 - 1 999', '2000-2999', '1 000 –1 999 бр'
    """
    if header is None:
        return None
    s = str(header).replace("–", "-").replace("—", "-")

    parts = re.findall(r"\d[\d\s\u00A0\u202F\.,]*\d|\d+", s)
    nums = []
    for p in parts:
        digits = re.sub(r"[^0-9]", "", p)
        if digits:
            try:
                nums.append(int(digits))
            except Exception:
                pass

    if len(nums) >= 2:
        return nums[0], nums[1]
    return None


def detect_range_columns(df_prices: pd.DataFrame):
    ranges = []
    for c in df_prices.columns:
        rng = parse_qty_range_from_header(c)
        if rng:
            ranges.append((rng[0], rng[1], c))
    ranges.sort(key=lambda x: (x[0], x[1]))
    return ranges


def resolve_unit_price_from_ranges(qty: int, price_row: pd.Series, ranges):
    if not ranges:
        return None

    for min_q, max_q, col in ranges:
        if min_q <= qty <= max_q:
            return to_float(price_row.get(col))

    # qty над последния диапазон -> ползвай последния
    last_min, last_max, last_col = ranges[-1]
    if qty > last_max:
        return to_float(price_row.get(last_col))

    return None


def find_column(df: pd.DataFrame, candidates):
    """
    Find a column by checking if any candidate substring appears in the header.
    Returns column name or raises ValueError.
    """
    cols = list(df.columns)
    norm_cols = {c: normalize(c) for c in cols}
    cand_norm = [normalize(x) for x in candidates]

    for c in cols:
        h = norm_cols[c]
        if any(cn in h for cn in cand_norm):
            return c
    raise ValueError(f"Не намерих колона за: {candidates}. Налични колони: {cols}")


def to_int(x):
    if pd.isna(x):
        return None
    if isinstance(x, (int,)):
        return int(x)
    try:
        s = str(x).strip()
        # remove thousands separators (spaces, NBSP)
        s = re.sub(r"[\s\u00A0\u202F]+", "", s)
        s = s.replace(",", ".")
        v = float(s)
        return int(round(v))
    except Exception:
        return None




def excel_cell_to_string(x) -> str:
    """Return value exactly as string from Excel (no parsing/formatting)."""
    if pd.isna(x):
        return ""
    return str(x).strip()


# -------------------------
# NEW: price ranges like "1 000 - 1 999"
# -------------------------
def parse_qty_range_from_header(header):
    """
    Extract (min,max) from headers like:
      "1 000 -1 999", "2 000 - 2 999", "1000-1999", "1 000 – 1 999"
      Also handles single numbers like 50000 -> (50000, 59999)
    Returns (min,max) or None.
    """
    if header is None:
        return None
    
    # If header is already a number
    if isinstance(header, (int, float)) and not pd.isna(header):
        qty = int(header)
        # Single number: treat as exact quantity with some range
        # e.g., 50000 means 50000-59999, 60000 means 60000-79999, etc.
        return (qty, qty)  # Will be handled specially in resolve_unit_price
    
    s = str(header).replace("–", "-").replace("—", "-")

    # find numbers possibly containing spaces/dots/commas (thousand separators)
    parts = re.findall(r"\d[\d\s\u00A0\u202F\.,]*\d|\d+", s)
    nums = []
    for p in parts:
        digits = re.sub(r"[^0-9]", "", p)
        if digits:
            try:
                nums.append(int(digits))
            except Exception:
                pass
    if len(nums) >= 2:
        return nums[0], nums[1]
    elif len(nums) == 1:
        # Single number in header text
        return (nums[0], nums[0])
    return None


def detect_range_columns(df_prices: pd.DataFrame):
    """
    Finds range-price columns in prices sheet and returns:
      [(min_qty, max_qty, colname), ...] sorted by min_qty.
    Also detects single-number columns like 50000, 60000.
    """
    ranges = []
    for c in df_prices.columns:
        rng = parse_qty_range_from_header(c)
        if rng:
            ranges.append((rng[0], rng[1], c))
    ranges.sort(key=lambda x: (x[0], x[1]))
    return ranges


def resolve_unit_price_from_ranges(qty: int, price_row: pd.Series, ranges):
    """
    Pick unit price by finding the column whose (min<=qty<=max).
    For single-number columns, picks the closest one that doesn't exceed qty.
    If qty is above the last range, use the last range's price (fallback).
    If qty is below the first range, use the first range's price (fallback).
    """
    if not ranges:
        return None

    # First, try exact range match
    for min_q, max_q, col in ranges:
        if min_q <= qty <= max_q:
            price = to_float(price_row.get(col))
            if price is not None:
                return price

    # For single-number columns (min==max), find the best match
    # Pick the highest single-number column that is <= qty
    single_cols = [(min_q, col) for min_q, max_q, col in ranges if min_q == max_q]
    if single_cols:
        # Sort by quantity descending
        single_cols.sort(key=lambda x: x[0], reverse=True)
        for threshold, col in single_cols:
            if qty >= threshold:
                price = to_float(price_row.get(col))
                if price is not None:
                    return price

    # fallback: qty above last range -> use last available price
    last_min, last_max, last_col = ranges[-1]
    if qty > last_max:
        price = to_float(price_row.get(last_col))
        if price is not None:
            return price

    # qty below first range -> use first available price
    first_min, first_max, first_col = ranges[0]
    if qty < first_min:
        price = to_float(price_row.get(first_col))
        if price is not None:
            return price
    
    # Try to find ANY non-null price in the ranges
    for min_q, max_q, col in ranges:
        price = to_float(price_row.get(col))
        if price is not None:
            return price

    return None


# -------------------------
# Excel reading
# -------------------------
def read_excel_any(path: str) -> pd.DataFrame:
    """
    Read .xls/.xlsx with pandas.
    .xls requires xlrd==2.0.1 installed.
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".xls":
            return pd.read_excel(path, engine="xlrd")
        else:
            return pd.read_excel(path, engine="openpyxl")
    except ImportError as e:
        raise ImportError(
            "Липсва библиотека за четене. Инсталирай:\n"
            "pip install pandas openpyxl xlrd==2.0.1\n\n"
            f"Оригинална грешка: {e}"
        )
    except Exception as e:
        raise RuntimeError(f"Не успях да прочета файла: {path}\nГрешка: {e}")


def merge_order_and_prices(order_path: str, prices_path: str) -> pd.DataFrame:
    df_order = read_excel_any(order_path)
    df_prices = read_excel_any(prices_path)

    DEBUG = os.environ.get("MERGE_DEBUG", "0") in ("1", "true", "True")
    if DEBUG:
        print("[DEBUG] Order columns:", list(df_order.columns))
        print("[DEBUG] Prices columns:", list(df_prices.columns))

    # Find order columns
    col_order_no = find_column(df_order, ["номер на поръчка", "поръчка", "Purchase Order"])
    col_item = find_column(df_order, ["име на артикул", "артикул", "продукт", "Item Number"])
    col_qty = find_column(df_order, ["заявени бройки", "бройки", "количество", "Quantity Ordered"])
    col_date = find_column(df_order, ["дата на доставка", "доставка", "delivery", "Due Date"])

    # Find prices columns
    p_item = find_column(df_prices, ["код АЛ филтър", "артикул", "item"])

    try:
        p_tl = find_column(df_prices, ["технологичен лист", "ТЛ", "tech"])
    except Exception:
        p_tl = None

    try:
        p_size = find_column(
            df_prices,
            ["размер", "шир./вис.", "шир/вис", "ширина/височина", "ширина/вис", "size", "width/height"]
        )
    except Exception:
        p_size = None
        # fallback: try to find a header containing both шир and вис
        for c in df_prices.columns:
            try:
                h = normalize(c)
                if "шир" in h and "вис" in h:
                    p_size = c
                    break
            except Exception:
                continue

    try:
        p_mat = find_column(df_prices, ["материал", "material"])
    except Exception:
        p_mat = None

    # Detect range columns like "1 000 - 1 999", "2 000 - 2 999", ...
    range_cols = detect_range_columns(df_prices)
    if not range_cols:
        raise ValueError(
            "Не открих ценови колони тип диапазон (напр. '1 000 - 1 999'). "
            "Провери заглавията в таблица 'Цени'."
        )

    # Build lookup: keep full row so we can read price from range columns
    # Use multiple keys for better matching: original, normalized, and canonical
    # When there are multiple rows for the same article, keep the one with more filled prices
    prices_lookup = {}
    
    def count_filled_prices(row):
        """Count how many price range columns have valid (non-NaN) values."""
        count = 0
        for _, _, col in range_cols:
            val = row.get(col)
            if not pd.isna(val):
                count += 1
        return count
    
    for _, pr in df_prices.iterrows():
        name = pr.get(p_item)
        if pd.isna(name):
            continue
        name = str(name).strip()
        name_norm = normalize(name)
        name_canon = canonical_code(name)
        info = {
            "row": pr,
            "Технологичен лист": "" if p_tl is None or pd.isna(pr.get(p_tl)) else str(pr.get(p_tl)).strip(),
            "Размер": "" if p_size is None or pd.isna(pr.get(p_size)) else str(pr.get(p_size)).strip(),
            "Материал": "" if p_mat is None or pd.isna(pr.get(p_mat)) else str(pr.get(p_mat)).strip(),
        }
        
        # Only update if this is a new key or if this row has more filled price columns
        new_count = count_filled_prices(pr)
        for key in [name, name_norm, name_canon]:
            if key is None:
                continue
            existing = prices_lookup.get(key)
            if existing is None:
                prices_lookup[key] = info
            else:
                # Compare: keep the row with more filled prices
                old_count = count_filled_prices(existing["row"])
                if new_count > old_count:
                    prices_lookup[key] = info

    if DEBUG:
        print(f"[DEBUG] Built prices_lookup with {len(prices_lookup)} keys. Sample keys: {list(prices_lookup.keys())[:6]}")

    # Merge
    out_rows = []
    line_no = 0

    for _, r in df_order.iterrows():
        order_no = r.get(col_order_no)
        item = r.get(col_item)
        qty = r.get(col_qty)
        ddate = r.get(col_date)

        if pd.isna(order_no) or pd.isna(item):
            continue

        order_no = str(order_no).strip()
        item = str(item).strip()
        item_norm = normalize(item)
        item_canon = canonical_code(item)

        qty_i = to_int(qty)
        if qty_i is None:
            continue

        line_no += 1
        order_ref = f"{order_no}-{line_no}"

        # find price row - STRICT matching by canonical code only
        # (артикулният номер от поръчката трябва да съвпада точно с "код АЛ филтър")
        price_info = None
        exact_match = False
        
        # Try exact canonical code match first
        if item_canon and item_canon in prices_lookup:
            price_info = prices_lookup.get(item_canon)
            exact_match = True
            if DEBUG:
                print(f"[DEBUG] Exact canonical match: '{item}' (canon='{item_canon}')")
        
        # Try exact normalized match
        if not exact_match and item_norm in prices_lookup:
            price_info = prices_lookup.get(item_norm)
            exact_match = True
            if DEBUG:
                print(f"[DEBUG] Exact normalized match: '{item}' (norm='{item_norm}')")
        
        # Try exact original match
        if not exact_match and item in prices_lookup:
            price_info = prices_lookup.get(item)
            exact_match = True
            if DEBUG:
                print(f"[DEBUG] Exact original match: '{item}'")
        
        # Try partial code match: if order item contains the price code or vice versa
        if not exact_match:
            for price_key, price_val in prices_lookup.items():
                if not isinstance(price_key, str) or not price_key:
                    continue
                price_canon = canonical_code(price_key)
                # Check if item code contains price code or price code contains item code
                if price_canon and item_canon:
                    if price_canon in item_canon or item_canon in price_canon:
                        price_info = price_val
                        exact_match = True
                        if DEBUG:
                            print(f"[DEBUG] Partial code match: '{item}' (canon='{item_canon}') contains/in '{price_key}' (canon='{price_canon}')")
                        break
        
        if not exact_match and DEBUG:
            print(f"[DEBUG] No exact match for '{item}' (canon='{item_canon}') - Ед. Цена and Сума will be empty")

        unit_price = None
        size = ""
        tl = ""
        mat = ""

        if price_info and exact_match:
            unit_price = resolve_unit_price_from_ranges(qty_i, price_info["row"], range_cols)
            size = price_info.get("Размер", "") or ""
            tl = price_info.get("Технологичен лист", "") or ""
            mat = price_info.get("Материал", "") or ""
            if DEBUG and unit_price is None:
                # show raw values in each range column for troubleshooting
                raw_vals = [(c, price_info["row"].get(c)) for _, _, c in range_cols]
                print(f"[DEBUG] Item '{item}' matched but unit_price=None. qty={qty_i}. Raw range values: {raw_vals[:6]}...")
        else:
            # No exact match - leave price fields empty
            if DEBUG:
                print(f"[DEBUG] Item '{item}' - no exact match, Ед. Цена and Сума will be empty.")

        total = round(unit_price * qty_i, 2) if unit_price is not None else None

        out_rows.append({
            "Артикул": item,
            "Размер": size,
            "Бройки": qty_i,
            "Ед. Цена": "" if unit_price is None else unit_price,
            "Сума": "" if total is None else total,
            "Номер на поръчка и ред": order_ref,
            "Дата на доставка": excel_cell_to_string(ddate),  # keep as string
            "Технологичен лист": tl,
            "Материал": mat,
        })

    return pd.DataFrame(out_rows)


def _apply_date_format_xlsx(path: Path, header_name: str = "Дата на доставка"):
    """
    Post-process an .xlsx file to ensure the column with header_name contains real datetimes
    and has a sensible Excel number format.
    NOTE: You wanted date as string, so this is typically NOT needed for your generated orders.
    It's kept only because your protocol logic uses it.
    """
    if load_workbook is None:
        return
    try:
        wb = load_workbook(filename=str(path))
        ws = wb.active
        header_col = None
        for cell in ws[1]:
            try:
                if str(cell.value).strip() == header_name:
                    header_col = cell.column
                    break
            except Exception:
                continue
        if header_col is None:
            wb.close()
            return

        for row in ws.iter_rows(min_row=2, min_col=header_col, max_col=header_col):
            cell = row[0]
            val = cell.value
            if val is None:
                continue
            import datetime as _dt
            if isinstance(val, (_dt.datetime, _dt.date)):
                cell.number_format = 'yyyy-mm-dd'
                continue

            try:
                parsed = pd.to_datetime(val, dayfirst=True, errors='coerce')
            except Exception:
                parsed = pd.to_datetime(val, errors='coerce')
            if pd.isna(parsed):
                continue
            try:
                cell.value = parsed.to_pydatetime()
            except Exception:
                continue
            cell.number_format = 'yyyy-mm-dd'

        wb.save(filename=str(path))
        wb.close()
    except Exception:
        try:
            wb.close()
        except Exception:
            pass


def load_settings():
    try:
        if SETTINGS_FILE.exists():
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def save_settings(s: dict):
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(s, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def set_protocols_dir(path_str: str):
    """Set protocols directory (no persistence - user must choose each session)."""
    global PROTOCOLS_DIR
    try:
        p = Path(path_str).expanduser().resolve()
    except Exception:
        p = Path(path_str)
    p.mkdir(parents=True, exist_ok=True)
    PROTOCOLS_DIR = p


def ensure_dirs():
    """Създава директориите за протоколи ако са избрани."""
    if PROTOCOLS_DIR is not None:
        try:
            PROTOCOLS_DIR.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass


def week_key_from_date(d):
    if pd.isna(d) or d == "":
        return "protocol_undated"
    if isinstance(d, str):
        s = d.strip()
        # Try ISO format first (YYYY-MM-DD) - must not use dayfirst for this
        if re.match(r"^\d{4}-\d{2}-\d{2}", s):
            dt = pd.to_datetime(s, errors="coerce")
        else:
            # For other formats like DD-MM-YYYY, DD/MM/YYYY, DD.MM.YYYY use dayfirst
            try:
                dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
            except Exception:
                dt = pd.to_datetime(s, errors="coerce")
    else:
        dt = pd.to_datetime(d, errors="coerce")

    if pd.isna(dt):
        return "protocol_undated"
    iso = dt.isocalendar()
    return f"protocol_{iso.year}_w{iso.week}"


def append_to_protocol(protocol_key: str, df_rows: pd.DataFrame, source_filename: str):
    ensure_dirs()
    prot_xlsx = PROTOCOLS_DIR / f"{protocol_key}.xlsx"
    
    # Check if protocol is closed (read-only or _CLOSED in name)
    if "_CLOSED" in protocol_key or is_file_readonly(prot_xlsx):
        raise RuntimeError(f"Протокол {protocol_key} е приключен. Нови редове не могат да се добавят.")

    cols = ["Артикул", "Размер", "Бройки", "Ед. Цена", "Сума", "Номер на поръчка и ред", "Дата на доставка", "Технологичен лист", "Материал"]
    out = df_rows.copy()
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    out = out[cols]

    if prot_xlsx.exists():
        try:
            existing = pd.read_excel(prot_xlsx, engine="openpyxl")
            
            # Remove duplicates: if "Номер на поръчка и ред" already exists, replace with new data
            if "Номер на поръчка и ред" in existing.columns and "Номер на поръчка и ред" in out.columns:
                # Get the order refs from new data
                new_refs = set(out["Номер на поръчка и ред"].dropna().astype(str).tolist())
                # Keep only rows from existing that are NOT in new data
                existing_filtered = existing[~existing["Номер на поръчка и ред"].astype(str).isin(new_refs)]
                new_all = pd.concat([existing_filtered, out], ignore_index=True)
            else:
                new_all = pd.concat([existing, out], ignore_index=True)
        except Exception:
            new_all = out
    else:
        new_all = out

    # If you really want protocols to have real date cells, keep this.
    if "Дата на доставка" in new_all.columns:
        try:
            new_all["Дата на доставка"] = pd.to_datetime(new_all["Дата на доставка"], errors="coerce")
        except Exception:
            pass

    new_all.to_excel(prot_xlsx, index=False)
    try:
        _apply_date_format_xlsx(prot_xlsx, header_name="Дата на доставка")
    except Exception:
        pass


# -------------------------
# Tkinter UI
# -------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Сливане на поръчка + цени (Excel)")
        self.geometry("1200x650")

        self.order_path = tk.StringVar(value="")
        self.prices_path = tk.StringVar(value="")
        self.protocols_dir_var = tk.StringVar(value="(не е избрана)")

        self.df_merged = None
        self._rendered_index_map = []
        self._current_file_path = None  # Path to currently loaded file for saving

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self, padding=10)
        top.pack(side=tk.TOP, fill=tk.X)

        btn_order = ttk.Button(top, text="Качи Поръчка (.xls/.xlsx)", command=self.pick_order)
        btn_prices = ttk.Button(top, text="Качи Цени (.xls/.xlsx)", command=self.pick_prices)
        btn_merge = ttk.Button(top, text="Слей", command=self.do_merge)
        btn_save = ttk.Button(top, text="Запази като...", command=self.save_xlsx)

        self.search_var = tk.StringVar(value="")
        self.search_entry = ttk.Entry(top, textvariable=self.search_var, width=30)
        btn_search = ttk.Button(top, text="Търси", command=self.on_search)

        # Row 1 buttons - single order processing
        btn_order.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        btn_prices.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        btn_merge.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        btn_save.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.search_entry.grid(row=0, column=4, padx=5, pady=5, sticky="w")
        btn_search.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        # Row 2 - protocol management buttons
        btn_choose_protocols = ttk.Button(top, text="Избери папка за протоколи", command=self.choose_protocols_folder)
        btn_batch = ttk.Button(top, text="Качи много поръчки", command=self.batch_process)
        btn_view_protocols = ttk.Button(top, text="Преглед протоколи", command=self.view_protocols)
        btn_close_protocol = ttk.Button(top, text="Приключи протокол", command=self.close_protocol)
        btn_reopen_protocol = ttk.Button(top, text="Отвори протокол", command=self.reopen_protocol)
        
        btn_choose_protocols.grid(row=1, column=0, padx=5, pady=2, sticky="w")
        btn_batch.grid(row=1, column=1, padx=5, pady=2, sticky="w")
        btn_view_protocols.grid(row=1, column=2, padx=5, pady=2, sticky="w")
        btn_close_protocol.grid(row=1, column=3, padx=5, pady=2, sticky="w")
        btn_reopen_protocol.grid(row=1, column=4, padx=5, pady=2, sticky="w")

        ttk.Label(top, text="Поръчка:").grid(row=2, column=0, sticky="w")
        ttk.Label(top, textvariable=self.order_path).grid(row=2, column=1, columnspan=6, sticky="w")

        ttk.Label(top, text="Цени:").grid(row=3, column=0, sticky="w")
        ttk.Label(top, textvariable=self.prices_path).grid(row=3, column=1, columnspan=6, sticky="w")

        ttk.Label(top, text="Протоколи: ").grid(row=4, column=0, sticky="w")
        ttk.Label(top, textvariable=self.protocols_dir_var).grid(row=4, column=1, columnspan=6, sticky="w")

        mid = ttk.Frame(self, padding=(10, 0, 10, 10))
        mid.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(mid, show="headings")
        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree.bind("<Double-1>", self.on_row_double_click)

        # Bottom bar with status
        self.status = tk.StringVar(value="Избери двата файла и натисни 'Слей'.")
        ttk.Label(self, textvariable=self.status, padding=10).pack(side=tk.BOTTOM, fill=tk.X)

    def pick_order(self):
        path = filedialog.askopenfilename(
            title="Избери файл Поръчка",
            filetypes=[("Excel", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("All files", "*.*")]
        )
        if path:
            self.order_path.set(path)

    def pick_prices(self):
        path = filedialog.askopenfilename(
            title="Избери файл Цени",
            filetypes=[("Excel", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("All files", "*.*")]
        )
        if path:
            self.prices_path.set(path)

    def choose_protocols_folder(self):
        path = filedialog.askdirectory(title="Избери папка за протоколи")
        if not path:
            return
        try:
            set_protocols_dir(path)
            self.protocols_dir_var.set(str(PROTOCOLS_DIR))
            self.status.set(f"Папка за протоколи: {PROTOCOLS_DIR}")
        except Exception as e:
            messagebox.showerror("Грешка", f"Не мога да задам папката: {e}")

    def view_protocols(self):
        """Show list of all protocols with their status."""
        if self.protocols_dir_var.get() == "(не е избрана)":
            messagebox.showwarning("Протоколи", "Първо избери папка за протоколи.")
            return
        
        protocols = []
        for p in PROTOCOLS_DIR.glob("protocol_*.xlsx"):
            name = p.stem
            # Check if closed by name (_CLOSED suffix)
            if "_CLOSED" in name:
                status = "ПРИКЛЮЧЕН"
                display_name = name.replace("_CLOSED", "")
            else:
                status = "Отворен"
                display_name = name
            try:
                df = pd.read_excel(p, engine="openpyxl")
                rows = len(df)
            except Exception:
                rows = "?"
            protocols.append((display_name, status, rows, name))
        
        if not protocols:
            messagebox.showinfo("Протоколи", "Няма намерени протоколи в избраната папка.")
            return
        
        # Create popup window
        popup = tk.Toplevel(self)
        popup.title("Протоколи")
        popup.geometry("600x400")
        
        tree = ttk.Treeview(popup, columns=("Име", "Статус", "Редове"), show="headings")
        tree.heading("Име", text="Протокол")
        tree.heading("Статус", text="Статус")
        tree.heading("Редове", text="Редове")
        tree.column("Име", width=300)
        tree.column("Статус", width=120)
        tree.column("Редове", width=80)
        
        for display_name, status, rows, full_name in sorted(protocols, reverse=True):
            tree.insert("", "end", values=(display_name, status, rows), tags=(full_name,))
        
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        def open_selected():
            sel = tree.selection()
            if not sel:
                return
            # Get full name from tags
            full_name = tree.item(sel[0])["tags"][0]
            prot_path = PROTOCOLS_DIR / f"{full_name}.xlsx"
            if prot_path.exists():
                try:
                    df = pd.read_excel(prot_path, engine="openpyxl")
                    self.df_merged = df
                    self._current_file_path = str(prot_path)  # Remember path for saving
                    self._load_table(df)
                    self.status.set(f"Заредени {len(df)} реда от {full_name} (двоен клик за редакция)")
                    popup.destroy()
                except Exception as e:
                    messagebox.showerror("Грешка", f"Не мога да отворя протокола: {e}")
        
        btn_open = ttk.Button(popup, text="Отвори в таблицата", command=open_selected)
        btn_open.pack(pady=5)

    def close_protocol(self):
        """Mark a protocol as closed (no more rows can be added)."""
        if self.protocols_dir_var.get() == "(не е избрана)":
            messagebox.showwarning("Протоколи", "Първо избери папка за протоколи.")
            return
        
        # Find protocols without _CLOSED in name
        open_protocols = [p.stem for p in PROTOCOLS_DIR.glob("protocol_*.xlsx") 
                          if "_CLOSED" not in p.stem]
        
        if not open_protocols:
            messagebox.showinfo("Протоколи", "Няма отворени протоколи за приключване.")
            return
        
        # Create selection dialog
        popup = tk.Toplevel(self)
        popup.title("Приключи протокол")
        popup.geometry("400x300")
        
        ttk.Label(popup, text="Избери протокол за приключване:").pack(pady=10)
        
        listbox = tk.Listbox(popup, selectmode=tk.SINGLE, width=50, height=10)
        for p in sorted(open_protocols, reverse=True):
            listbox.insert(tk.END, p)
        listbox.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        
        def do_close():
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("Избор", "Избери протокол от списъка.")
                return
            name = listbox.get(sel[0])
            
            if messagebox.askyesno("Потвърждение", 
                f"Сигурен ли си, че искаш да приключиш протокол '{name}'?\n\n"
                "След приключване няма да можеш да добавяш нови редове към него.\n"
                "Файлът ще бъде преименуван с _CLOSED и защитен от промени."):
                
                prot_file = PROTOCOLS_DIR / f"{name}.xlsx"
                closed_file = PROTOCOLS_DIR / f"{name}_CLOSED.xlsx"
                
                if prot_file.exists():
                    # Rename to _CLOSED
                    try:
                        prot_file.rename(closed_file)
                        # Set file as read-only
                        set_file_readonly(closed_file, readonly=True)
                        messagebox.showinfo("Готово", f"Протокол '{name}' е приключен.\nНово име: {closed_file.name}")
                    except Exception as e:
                        messagebox.showerror("Грешка", f"Не мога да преименувам файла: {e}")
                        return
                
                popup.destroy()
        
        ttk.Button(popup, text="Приключи", command=do_close).pack(pady=10)

    def reopen_protocol(self):
        """Reopen a closed protocol."""
        if self.protocols_dir_var.get() == "(не е избрана)":
            messagebox.showwarning("Протоколи", "Първо избери папка за протоколи.")
            return
        
        # Find protocols with _CLOSED in name
        closed_protocols = [p.stem for p in PROTOCOLS_DIR.glob("protocol_*_CLOSED.xlsx")]
        
        if not closed_protocols:
            messagebox.showinfo("Протоколи", "Няма приключени протоколи за отваряне.")
            return
        
        # Create selection dialog
        popup = tk.Toplevel(self)
        popup.title("Отвори протокол")
        popup.geometry("400x300")
        
        ttk.Label(popup, text="Избери протокол за отваряне:").pack(pady=10)
        
        listbox = tk.Listbox(popup, selectmode=tk.SINGLE, width=50, height=10)
        for p in sorted(closed_protocols, reverse=True):
            listbox.insert(tk.END, p)
        listbox.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        
        def do_reopen():
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("Избор", "Избери протокол от списъка.")
                return
            name = listbox.get(sel[0])  # e.g. protocol_2026_w7_CLOSED
            
            closed_file = PROTOCOLS_DIR / f"{name}.xlsx"
            # Remove _CLOSED from name
            open_name = name.replace("_CLOSED", "")
            open_file = PROTOCOLS_DIR / f"{open_name}.xlsx"
            
            if closed_file.exists():
                try:
                    # Remove read-only flag first
                    set_file_readonly(closed_file, readonly=False)
                    # Rename back
                    closed_file.rename(open_file)
                    messagebox.showinfo("Готово", f"Протокол '{open_name}' е отворен за добавяне на редове.")
                except Exception as e:
                    messagebox.showerror("Грешка", f"Не мога да преименувам файла: {e}")
                    return
            
            popup.destroy()
        
        ttk.Button(popup, text="Отвори", command=do_reopen).pack(pady=10)

    def batch_process(self):
        """Process multiple order files at once and add them to weekly protocols."""
        if self.protocols_dir_var.get() == "(не е избрана)" or PROTOCOLS_DIR is None:
            messagebox.showwarning("Папка за протоколи", "Първо избери папка за протоколи.")
            return
        
        # Create window
        popup = tk.Toplevel(self)
        popup.title("Качи много поръчки")
        popup.geometry("600x500")
        popup.transient(self)
        
        order_files = []
        
        # Title
        ttk.Label(popup, text="Качи поръчки към протоколи", font=("", 14, "bold")).pack(pady=(15, 10))
        
        # Add files button - PROMINENT at top
        def add_files_dialog():
            paths = filedialog.askopenfilenames(
                title="Избери поръчки",
                filetypes=[("Excel", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("All files", "*.*")]
            )
            if paths:
                for p in paths:
                    p_str = str(p).strip()
                    if p_str.lower().endswith(('.xls', '.xlsx')) and p_str not in order_files:
                        order_files.append(p_str)
                        files_listbox.insert(tk.END, Path(p_str).name)
                update_label()
        
        add_btn = ttk.Button(popup, text="📂 Добави файлове...", command=add_files_dialog)
        add_btn.pack(pady=10)
        
        # Files list label
        ttk.Label(popup, text="Добавени поръчки:").pack(anchor="w", padx=20)
        
        # Listbox for files
        list_frame = ttk.Frame(popup)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        files_listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED, height=12)
        files_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=files_listbox.yview)
        files_listbox.configure(yscrollcommand=files_scrollbar.set)
        
        files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        files_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Status label
        status_label = ttk.Label(popup, text="Няма добавени файлове", foreground="gray")
        status_label.pack(pady=5)
        
        def update_label():
            if order_files:
                status_label.configure(text=f"{len(order_files)} файла добавени", foreground="green")
            else:
                status_label.configure(text="Няма добавени файлове", foreground="gray")
        
        # Action buttons
        btn_frame = ttk.Frame(popup)
        btn_frame.pack(fill=tk.X, padx=20, pady=5)
        
        def remove_selected():
            selected = list(files_listbox.curselection())
            for i in reversed(selected):
                files_listbox.delete(i)
                del order_files[i]
            update_label()
        
        def clear_all():
            files_listbox.delete(0, tk.END)
            order_files.clear()
            update_label()
        
        ttk.Button(btn_frame, text="Премахни избраните", command=remove_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Изчисти всички", command=clear_all).pack(side=tk.LEFT, padx=5)
        
        # Separator and bottom buttons
        ttk.Separator(popup, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=20, pady=10)
        
        bottom_frame = ttk.Frame(popup)
        bottom_frame.pack(fill=tk.X, padx=20, pady=10)
        
        def do_process():
            if not order_files:
                messagebox.showwarning("Липсват файлове", "Добави поне една поръчка.")
                return
            popup.destroy()
            self._process_batch_files(order_files)
        
        ttk.Button(bottom_frame, text="Отказ", command=popup.destroy).pack(side=tk.RIGHT, padx=5)
        ttk.Button(bottom_frame, text="Обработи", command=do_process).pack(side=tk.RIGHT, padx=5)

    def _process_batch_files(self, order_files):
        """Process the batch of order files (already contain prices)."""
        processed = 0
        errors = []
        all_merged = []
        
        # Expected columns in protocol
        protocol_cols = ["Артикул", "Размер", "Бройки", "Ед. Цена", "Сума", 
                         "Номер на поръчка и ред", "Дата на доставка", "Технологичен лист", "Материал"]
        
        # Column name mappings (various formats -> standard format)
        col_mappings = {
            # Артикул
            "артикул": "Артикул",
            "item": "Артикул",
            "item number": "Артикул",
            "име на артикул": "Артикул",
            "продукт": "Артикул",
            # Размер
            "размер": "Размер",
            "size": "Размер",
            # Бройки
            "бройки": "Бройки",
            "брой": "Бройки",
            "qty": "Бройки",
            "quantity": "Бройки",
            "количество": "Бройки",
            # Ед. Цена
            "ед. цена": "Ед. Цена",
            "ед цена": "Ед. Цена",
            "единична цена": "Ед. Цена",
            "unit price": "Ед. Цена",
            "цена": "Ед. Цена",
            # Сума
            "сума": "Сума",
            "total": "Сума",
            "amount": "Сума",
            "обща сума": "Сума",
            # Номер на поръчка и ред
            "номер на поръчка и ред": "Номер на поръчка и ред",
            "номер поръчка": "Номер на поръчка и ред",
            "поръчка": "Номер на поръчка и ред",
            "order": "Номер на поръчка и ред",
            "order number": "Номер на поръчка и ред",
            "purchase order": "Номер на поръчка и ред",
            # Дата на доставка
            "дата на доставка": "Дата на доставка",
            "дата доставка": "Дата на доставка",
            "дата": "Дата на доставка",
            "delivery date": "Дата на доставка",
            "date": "Дата на доставка",
            # Технологичен лист
            "технологичен лист": "Технологичен лист",
            "тл": "Технологичен лист",
            "tech sheet": "Технологичен лист",
            # Материал
            "материал": "Материал",
            "material": "Материал",
        }
        
        # Required columns - file must have at least these
        required_cols = ["Артикул", "Бройки"]
        
        for order_path in order_files:
            try:
                # Read the order file directly (it already has all data including prices)
                df = read_excel_any(order_path)
                
                if df.empty:
                    errors.append(f"{Path(order_path).name}: Файлът е празен")
                    continue
                
                # Normalize column names
                new_columns = {}
                for col in df.columns:
                    col_lower = str(col).strip().lower()
                    if col_lower in col_mappings:
                        new_columns[col] = col_mappings[col_lower]
                    elif str(col).strip() in protocol_cols:
                        new_columns[col] = str(col).strip()
                
                if new_columns:
                    df = df.rename(columns=new_columns)
                
                # Check if file has required columns
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    errors.append(f"{Path(order_path).name}: Липсват колони: {', '.join(missing_cols)}")
                    continue
                
                # Get order name from filename
                order_no = Path(order_path).stem
                
                # Ensure all protocol columns exist (add missing optional ones as empty)
                for col in protocol_cols:
                    if col not in df.columns:
                        df[col] = ""
                
                # Group by week and append to protocols
                ensure_dirs()
                groups = {}
                for _, row in df.iterrows():
                    wk = week_key_from_date(row.get("Дата на доставка"))
                    groups.setdefault(wk, []).append(row.to_dict())
                
                for wk, rows in groups.items():
                    df_rows = pd.DataFrame(rows)
                    try:
                        append_to_protocol(wk, df_rows, Path(order_path).name)
                    except RuntimeError as e:
                        # Protocol is closed
                        errors.append(f"{order_no}: {e}")
                
                processed += 1
                all_merged.append(df)
                
            except Exception as e:
                errors.append(f"{Path(order_path).name}: {e}")
        
        # Show results
        if all_merged:
            combined = pd.concat(all_merged, ignore_index=True)
            self.df_merged = combined
            self._load_table(combined)
        
        msg = f"Обработени {processed} от {len(order_files)} поръчки."
        if errors:
            msg += f"\n\nГрешки:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                msg += f"\n... и още {len(errors) - 10} грешки"
            messagebox.showwarning("Резултат", msg)
        else:
            messagebox.showinfo("Готово", msg)
        
        self.status.set(msg.split("\n")[0])

    def do_merge(self):
        op = self.order_path.get().strip()
        pp = self.prices_path.get().strip()
        if not op or not pp:
            messagebox.showwarning("Липсват файлове", "Моля избери и двата файла (Поръчка и Цени).")
            return

        try:
            self.df_merged = merge_order_and_prices(op, pp)
            self._current_file_path = None  # New merge, no file yet
            self._load_table(self.df_merged)
            self.status.set(f"Готово: {len(self.df_merged)} реда слети.")
        except Exception as e:
            messagebox.showerror("Грешка", str(e))
            self.status.set("Грешка при сливане.")

    def save_xlsx(self):
        if self.df_merged is None or self.df_merged.empty:
            messagebox.showinfo("Няма данни", "Първо натисни 'Слей'.")
            return

        default_name = "Porachka.xlsx"
        try:
            first_ref = str(self.df_merged.iloc[0]["Номер на поръчка и ред"])
            order_no = first_ref.split("-")[0]
            default_name = f"Porachka_{order_no}.xlsx"
        except Exception:
            pass

        out_path = filedialog.asksaveasfilename(
            title="Запази като",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")]
        )
        if not out_path:
            return

        try:
            # Преобразуваме датата в реална дата за Numbers/Excel съвместимост
            df_to_save = self.df_merged.copy()
            if "Дата на доставка" in df_to_save.columns:
                df_to_save["Дата на доставка"] = pd.to_datetime(df_to_save["Дата на доставка"], errors="coerce")
            
            # Ако файлът съществува, презаписваме дублиращите се редове
            if Path(out_path).exists():
                try:
                    existing = pd.read_excel(out_path, engine="openpyxl")
                    if "Номер на поръчка и ред" in existing.columns and "Номер на поръчка и ред" in df_to_save.columns:
                        # Get the order refs from new data
                        new_refs = set(df_to_save["Номер на поръчка и ред"].dropna().astype(str).tolist())
                        # Keep only rows from existing that are NOT in new data
                        existing_filtered = existing[~existing["Номер на поръчка и ред"].astype(str).isin(new_refs)]
                        df_to_save = pd.concat([existing_filtered, df_to_save], ignore_index=True)
                        # Convert date again after merge
                        if "Дата на доставка" in df_to_save.columns:
                            df_to_save["Дата на доставка"] = pd.to_datetime(df_to_save["Дата на доставка"], errors="coerce")
                except Exception:
                    pass  # If can't read existing, just overwrite
            
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df_to_save.to_excel(writer, index=False, sheet_name="Porachka")
            
            # Прилагаме формат за дата
            try:
                _apply_date_format_xlsx(out_path, header_name="Дата на доставка")
            except Exception:
                pass
            
            messagebox.showinfo("Записано", f"Файлът е записан:\n{out_path}")
            self.status.set(f"Записано: {out_path}")
        except Exception as e:
            messagebox.showerror("Грешка при запис", str(e))
            self.status.set("Грешка при запис.")
            return

        # Check if protocols folder is selected - append to weekly protocols
        if self.protocols_dir_var.get() == "(не е избрана)":
            # No protocols folder selected, skip adding to protocols
            return

        # Append to weekly protocols
        try:
            ensure_dirs()
            source_name = Path(out_path).name

            # append each row to its weekly protocol
            groups = {}
            for _, row in self.df_merged.iterrows():
                wk = week_key_from_date(row.get("Дата на доставка"))
                groups.setdefault(wk, []).append(row.to_dict())

            for wk, rows in groups.items():
                df_rows = pd.DataFrame(rows)
                try:
                    append_to_protocol(wk, df_rows, source_name)
                except Exception as e:
                    messagebox.showwarning("Протокол", str(e))

        except Exception as e:
            messagebox.showwarning("Добавяне към протокол", f"Грешка: {e}")

    def _load_table(self, df: pd.DataFrame):
        for col in self.tree["columns"]:
            self.tree.heading(col, text="")
        self.tree.delete(*self.tree.get_children())

        cols = list(df.columns)
        self.tree["columns"] = cols

        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=140, anchor="w")

        max_rows = 2000
        self._rendered_index_map = []
        idx_counter = 0

        for i, row in df.head(max_rows).iterrows():
            values = []
            for c in cols:
                v = row.get(c, "")
                try:
                    if pd.isna(v):
                        values.append("")
                        continue
                except Exception:
                    pass
                values.append(str(v))
            iid = str(idx_counter)
            self.tree.insert("", "end", iid=iid, values=values)
            self._rendered_index_map.append(i)
            idx_counter += 1

        if len(df) > max_rows:
            self.status.set(f"Показвам първите {max_rows} реда от {len(df)} (всички се записват при export).")

    def on_search(self):
        q = (self.search_var.get() or "").strip()
        if not q:
            messagebox.showinfo("Търсене", "Въведи текст за търсене (име, ТЛ или размер).")
            return
        ql = q.lower()

        def match_row(r):
            for c in ["Артикул", "Технологичен лист", "Размер"]:
                try:
                    v = str(r.get(c, "") or "").lower()
                except Exception:
                    v = ""
                if ql in v:
                    return True
            return False

        if getattr(self, 'df_merged', None) is not None and not self.df_merged.empty:
            try:
                filtered = self.df_merged[self.df_merged.apply(match_row, axis=1)]
                if not filtered.empty:
                    self.df_merged = filtered
                    self._load_table(self.df_merged)
                    self.status.set(f"Намерени {len(filtered)} реда за '{q}' (в текущото сливане).")
                    return
            except Exception:
                pass

        # No local orders directory - just search in current merge
        messagebox.showinfo("Търсене", "Няма текущо сливане за търсене. Първо заредете поръчка.")

    def on_row_double_click(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        try:
            idx = int(iid)
        except Exception:
            return
        if idx >= len(self._rendered_index_map):
            return
        df_index = self._rendered_index_map[idx]

        columns = list(self.tree['columns'])
        cur_values = [self.df_merged.at[df_index, c] if c in self.df_merged.columns else '' for c in columns]

        edit = tk.Toplevel(self)
        edit.title("Редакция на ред")

        entries = {}
        for i, c in enumerate(columns):
            ttk.Label(edit, text=c).grid(row=i, column=0, sticky='w', padx=4, pady=2)
            # Handle NaN/None values - show empty string instead of "nan"
            val = cur_values[i]
            if val is None or (isinstance(val, float) and pd.isna(val)):
                display_val = ""
            else:
                try:
                    if pd.isna(val):
                        display_val = ""
                    else:
                        display_val = str(val)
                except Exception:
                    display_val = str(val) if val is not None else ""
            v = tk.StringVar(value=display_val)
            e = ttk.Entry(edit, textvariable=v, width=60)
            e.grid(row=i, column=1, sticky='w', padx=4, pady=2)
            entries[c] = v

        def save_edit():
            for c, var in entries.items():
                val = var.get().strip()
                
                # Handle empty values
                if val == "" or val.lower() == "nan":
                    if c in ("Бройки",):
                        self.df_merged.at[df_index, c] = None
                    elif c in ("Ед. Цена", "Сума", "Технологичен лист"):
                        self.df_merged.at[df_index, c] = None
                    else:
                        self.df_merged.at[df_index, c] = ""
                elif c in ("Бройки",):
                    try:
                        vv = int(val)
                        self.df_merged.at[df_index, c] = vv
                    except Exception:
                        self.df_merged.at[df_index, c] = None
                elif c in ("Ед. Цена", "Сума", "Технологичен лист"):
                    try:
                        vv = float(str(val).replace(',', '.'))
                        self.df_merged.at[df_index, c] = vv
                    except Exception:
                        self.df_merged.at[df_index, c] = None
                else:
                    self.df_merged.at[df_index, c] = val

            edit.destroy()
            self._load_table(self.df_merged)
            
            # Auto-save to file if we have a current file path
            if self._current_file_path:
                try:
                    out_path = self._current_file_path
                    
                    # Check if file is read-only (closed protocol)
                    if is_file_readonly(Path(out_path)):
                        self.status.set("⚠️ Файлът е защитен - промените не са запазени.")
                        return
                    
                    # Save to file
                    df_to_save = self.df_merged.copy()
                    if "Дата на доставка" in df_to_save.columns:
                        df_to_save["Дата на доставка"] = pd.to_datetime(df_to_save["Дата на доставка"], errors="coerce")
                    
                    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                        df_to_save.to_excel(writer, index=False, sheet_name="Sheet1")
                    
                    try:
                        _apply_date_format_xlsx(out_path, header_name="Дата на доставка")
                    except Exception:
                        pass
                    
                    self.status.set(f"✅ Запазено в {Path(out_path).name}")
                except Exception as e:
                    self.status.set(f"⚠️ Грешка при запис: {e}")
            else:
                self.status.set("Редът е променен (използвай 'Запази като...' за да запишеш)")

        btn_save = ttk.Button(edit, text="Запиши", command=save_edit)
        btn_save.grid(row=len(columns), column=0, columnspan=2, pady=6)


if __name__ == "__main__":
    # IMPORTANT: Windows-only DPI tweak; do NOT run on macOS/Linux
    if os.name == "nt":
        try:
            import ctypes
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

    app = App()
    app.mainloop()
